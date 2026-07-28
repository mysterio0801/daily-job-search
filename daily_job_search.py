#!/usr/bin/env python3
"""
Daily job-match pipeline.

Collect -> hard-filter -> LLM-score -> ranked .xlsx

Env vars required:
    APIFY_TOKEN         Apify API token (Settings -> Integrations -> API tokens)
    ANTHROPIC_API_KEY   Anthropic API key

Config:
    config/profile.local.yaml   your profile/rubric/search/referrals (gitignored)
    config/profile.example.yaml fallback demo profile, used if the above is absent

Usage:
    python daily_job_search.py                 # normal daily run
    python daily_job_search.py --window 48h    # widen the posting window
    python daily_job_search.py --dry-run       # collect + filter, skip scoring (free)
    python daily_job_search.py --config path/to/profile.yaml
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime, timezone

import requests
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

APIFY_ACTOR = "cheap_scraper~linkedin-job-scraper"
APIFY_SYNC = f"https://api.apify.com/v2/acts/{APIFY_ACTOR}/run-sync-get-dataset-items"
ANTHROPIC_URL = "https://api.anthropic.com/v1/messages"
MODEL = os.environ.get("JOBSEARCH_MODEL", "claude-sonnet-5")

LOCAL_CONFIG_PATH = "config/profile.local.yaml"
EXAMPLE_CONFIG_PATH = "config/profile.example.yaml"

# --------------------------------------------------------------------------
# Config
# --------------------------------------------------------------------------

def load_config(path: str | None) -> dict:
    """Load profile/rubric/search/referrals config.

    Falls back to the tracked example profile (a fictional demo persona) if
    profile.local.yaml hasn't been created yet, so the pipeline runs out of
    the box without requiring personal data to exist.
    """
    chosen = path
    if not chosen:
        if os.path.exists(LOCAL_CONFIG_PATH):
            chosen = LOCAL_CONFIG_PATH
        else:
            chosen = EXAMPLE_CONFIG_PATH
            print(
                f"[config] {LOCAL_CONFIG_PATH} not found -- using {EXAMPLE_CONFIG_PATH} "
                f"(fictional demo profile). Copy it to {LOCAL_CONFIG_PATH} and fill in "
                f"your own details.",
                file=sys.stderr,
            )
    with open(chosen) as f:
        cfg = yaml.safe_load(f)
    cfg.setdefault("referrals", {})
    cfg.setdefault("tiers", {})
    cfg["tiers"].setdefault("target_rows", 20)
    cfg["tiers"].setdefault("tier1_bar", 80)
    cfg["tiers"].setdefault("tier2_bar", 70)
    cfg.setdefault("search", {})
    cfg["search"].setdefault("keywords", ["Software Engineer"])
    cfg["search"].setdefault("passes", [{"locations": [], "remote": False}])
    cfg["search"].setdefault("exclude_companies", [])
    return cfg


# --------------------------------------------------------------------------
# Hard filters (generic engine logic -- shared across every profile/fork)
# --------------------------------------------------------------------------

TITLE_REJECT = re.compile(
    r"\b(senior|sr\.?|staff|principal|lead|manager|director|architect|intern|"
    r"fresher|sdet|qa|tester|testing|android|ios|mobile|flutter|react native|"
    r"frontend|front.end|ui developer|devops|sre|site reliability|"
    r"data scientist|data science|embedded|firmware|\.net|golang|rust|"
    r"abap|sap|pega|mulesoft|guidewire|appian|mendix|windchill|sccm|"
    r"walk.?in|consultant|instructor|mentor|freelance|vendor)\b",
    re.I,
)

# Base denylist: staffing aggregators / IT-services & consulting firms. Generic
# across profiles; a fork adds its own extras via search.exclude_companies.
COMPANY_REJECT = {
    "scoutit", "duruper", "studyecart technologies pvt ltd", "crossing hurdles",
    "impetus career consultants", "applicantz", "talentquell", "jobgether",
    "hackajob", "hired", "avua", "tekpillar", "themesoft inc.", "anlage digital",
    "golden opportunities", "recro", "qualizeal",
    "infosys", "tata consultancy services", "wipro", "accenture in india",
    "cognizant", "capgemini", "hcltech", "virtusa", "ust", "lti", "ltimindtree",
    "tech mahindra", "birlasoft", "synechron", "ntt data north america",
    "cyient", "codevyasa", "indium software", "ascendion", "happiest minds technologies",
    "quest global", "eclerx", "infobeans", "tata technologies",
}

# Language signals. A JD naming none of the ACCEPT terms, or naming a REJECT
# term without any ACCEPT term, is not your stack.
LANG_ACCEPT = re.compile(r"\b(java|kotlin|jvm|spring|ktor|scala)\b", re.I)
LANG_REJECT = re.compile(
    r"\b(c\+\+|\.net|c#|golang|\bgo\b|rust|php|ruby|rails|perl|cobol|pl/sql)\b", re.I
)

# Discipline signals in the BODY. Necessary because a title filter cannot catch
# an Android role -- Android JDs name Kotlin, so LANG_ACCEPT happily passes them.
JD_REJECT = re.compile(
    r"(jetpack compose|android sdk|androidx|android development|native android|"
    r"swiftui|\bxcode\b|\bios development\b|react native|\bflutter\b|"
    r"mobile app development|mobile application development|"
    r"\bllm\b|prompt engineering|agentic|\brag\b|langchain|langgraph|"
    r"fine.?tuning|model training|pytorch|tensorflow)",
    re.I,
)

POSTED_HOURS = re.compile(r"(\d+)\s*(minute|hour|day)", re.I)


def posted_hours_ago(posted: str) -> float:
    """'3 hours ago' -> 3.0 ; '28 minutes ago' -> 0.47 ; '1 day ago' -> 24.0"""
    m = POSTED_HOURS.search(posted or "")
    if not m:
        return 999.0
    n, unit = int(m.group(1)), m.group(2).lower()
    return {"minute": n / 60, "hour": float(n), "day": n * 24.0}[unit]


def hard_filter(job: dict, max_hours: float, exclude_companies: set) -> str | None:
    """Return a rejection reason, or None if the job survives."""
    title = job.get("jobTitle") or ""
    company = (job.get("companyName") or "").strip().lower()
    jd = job.get("jobDescription") or ""

    if posted_hours_ago(job.get("postedTime", "")) > max_hours:
        return "outside posting window"
    if TITLE_REJECT.search(title):
        return "title excluded"
    if company in COMPANY_REJECT or company in exclude_companies:
        return "company excluded"
    if len(jd) < 400:
        return "description too thin to score"
    if not LANG_ACCEPT.search(jd):
        return "no JVM language in description"
    if LANG_REJECT.search(jd) and len(LANG_ACCEPT.findall(jd)) < 2:
        return "primary language is not JVM"
    hits = JD_REJECT.findall(jd)
    if len(hits) >= 2:
        return f"wrong discipline (description signals: {hits[0]})"

    years = (job.get("yearsOfExperience") or {}).get("years")
    if isinstance(years, (int, float)) and years > 5:
        return f"requires {years}+ years"
    return None


# --------------------------------------------------------------------------
# Collection
# --------------------------------------------------------------------------

def apify_input(keywords, locations, window, remote=False) -> dict:
    payload = {
        "keyword": keywords,
        "locations": locations,
        "publishedAt": window,
        "jobType": ["full-time"],
        "experienceLevel": ["associate", "mid-senior"],
        "excludeRecruitingAgencies": True,
        "saveOnlyUniqueItems": True,
        "enrichCompanyData": True,
        "maxItems": 200,
    }
    if remote:
        payload["workType"] = ["remote"]
    return payload


def collect(token: str, window: str, search_cfg: dict) -> list[dict]:
    jobs, seen = [], set()
    keywords = search_cfg["keywords"]
    for i, p in enumerate(search_cfg["passes"]):
        name = p.get("name", f"pass-{i + 1}")
        locations = p["locations"]
        remote = p.get("remote", False)
        print(f"[collect] {name} ...", flush=True)
        try:
            r = requests.post(
                APIFY_SYNC,
                params={"token": token, "memory": 1024, "timeout": 300},
                json=apify_input(keywords, locations, window, remote),
                timeout=420,
            )
            r.raise_for_status()
            batch = r.json()
        except Exception as e:
            # An unattended run has nobody to retry it. Log loudly, keep going.
            print(f"[collect] {name} FAILED: {e}", file=sys.stderr)
            continue
        new = 0
        for j in batch:
            jid = j.get("jobId")
            if jid and jid not in seen:
                seen.add(jid)
                jobs.append(j)
                new += 1
        print(f"[collect] {name}: {len(batch)} returned, {new} new", flush=True)
    return jobs


# --------------------------------------------------------------------------
# Scoring
# --------------------------------------------------------------------------

def build_scoring_system(profile: str, rubric: str) -> str:
    return f"""You score job descriptions against one candidate profile.

CANDIDATE PROFILE
{profile}

RUBRIC
{rubric}

Score from the DESCRIPTION TEXT, never the title. Titles routinely lie: postings
titled "Software Dev Engineer II" have turned out to be Native Android roles and
Python AI roles; a "Software Engineer" posting was C/C++ routing protocols.

Be strict. A JD that merely lists a language among ten in a boilerplate skills
block does not earn full core_stack marks. Reserve 30+ on core_stack for JDs where
the target stack is clearly the primary stack of the role.

Respond with ONLY a JSON array, no prose and no markdown fences. One object per job,
in the same order given:
[{{"i":0,"score":83,"core_stack":28,"system_type":23,"cloud_infra":12,
   "seniority":10,"location":10,"ai_bonus":0,
   "why":"one sentence on what earned or lost the points",
   "caveat":"the single biggest reason to hesitate, or empty string"}}]
"""


def score_batch(api_key: str, system_prompt: str, batch: list[dict]) -> list[dict]:
    payload = []
    for i, j in enumerate(batch):
        payload.append(
            {
                "i": i,
                "title": j.get("jobTitle"),
                "company": j.get("companyName"),
                "location": j.get("location"),
                "years": (j.get("yearsOfExperience") or {}).get("years"),
                "description": (j.get("jobDescription") or "")[:6000],
            }
        )
    r = requests.post(
        ANTHROPIC_URL,
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json={
            "model": MODEL,
            "max_tokens": 4000,
            "system": system_prompt,
            "messages": [{"role": "user", "content": json.dumps(payload)}],
        },
        timeout=300,
    )
    r.raise_for_status()
    text = "".join(b.get("text", "") for b in r.json()["content"] if b["type"] == "text")
    text = re.sub(r"^```(?:json)?|```$", "", text.strip(), flags=re.M).strip()
    return json.loads(text)


def score_all(
    api_key: str, system_prompt: str, jobs: list[dict], batch_size: int = 6
) -> list[dict]:
    scored = []
    for start in range(0, len(jobs), batch_size):
        batch = jobs[start : start + batch_size]
        print(f"[score] {start + 1}-{start + len(batch)} of {len(jobs)} ...", flush=True)
        try:
            results = score_batch(api_key, system_prompt, batch)
        except Exception as e:
            print(f"[score] batch failed, marking unverified: {e}", file=sys.stderr)
            results = [{"i": i, "score": None} for i in range(len(batch))]
        by_i = {r["i"]: r for r in results if isinstance(r, dict) and "i" in r}
        for i, job in enumerate(batch):
            res = by_i.get(i, {"score": None})
            job["_score"] = res.get("score")
            job["_why"] = res.get("why", "")
            job["_caveat"] = res.get("caveat", "")
            scored.append(job)
    return scored


# --------------------------------------------------------------------------
# Output
# --------------------------------------------------------------------------

HEADERS = [
    "#", "Tier", "Score", "Role", "Company", "Location", "Posted", "Applicants",
    "Apply Via", "Job URL", "Why it scored", "Caveat", "Referral path",
    "Status", "Date applied", "Notes",
]


def tier_of(score, tier1_bar: int, tier2_bar: int) -> str:
    if score is None:
        return "Tier 3"
    if score >= tier1_bar:
        return "Tier 1"
    if score >= tier2_bar:
        return "Tier 2"
    return "Below bar"


def write_xlsx(
    jobs: list[dict], path: str, referrals: dict, target_rows: int,
    tier1_bar: int, tier2_bar: int,
) -> tuple[int, int]:
    ranked = [
        j for j in jobs
        if tier_of(j.get("_score"), tier1_bar, tier2_bar) != "Below bar"
    ]
    ranked.sort(key=lambda j: posted_hours_ago(j.get("postedTime", "")))
    ranked = ranked[:target_rows]
    tier1 = sum(
        1 for j in ranked if tier_of(j["_score"], tier1_bar, tier2_bar) == "Tier 1"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Job Matches"
    ws.append(HEADERS)

    fills = {
        "Tier 1": PatternFill("solid", fgColor="D9EAD3"),
        "Tier 2": PatternFill("solid", fgColor="FFF2CC"),
        "Tier 3": PatternFill("solid", fgColor="F2F2F2"),
    }
    edit_fill = PatternFill("solid", fgColor="FFFF00")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for n, j in enumerate(ranked, start=1):
        tier = tier_of(j["_score"], tier1_bar, tier2_bar)
        company = (j.get("companyName") or "").strip()
        ws.append([
            n, tier, j["_score"] if j["_score"] is not None else "NOT SCORED",
            j.get("jobTitle"), company, j.get("location"), j.get("postedTime"),
            j.get("applicationsCount") or "-",
            "Easy Apply" if j.get("applyType") == "EASY_APPLY" else "External",
            j.get("jobUrl"),
            j.get("_why") or "NOT JD-VERIFIED - scoring failed for this row",
            j.get("_caveat") or "",
            referrals.get(company.lower(), "No warm contact on file"),
            "", "", "",
        ])

    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    for r in range(2, len(ranked) + 2):
        tier = ws.cell(row=r, column=2).value
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.fill = edit_fill if c >= 14 else fills[tier]
        link = ws.cell(row=r, column=10)
        if link.value:
            link.hyperlink = link.value
            link.font = Font(name="Arial", size=9, color="0563C1", underline="single")
        ws.row_dimensions[r].height = 56

    for i, w in enumerate([4, 8, 7, 34, 22, 20, 12, 11, 12, 44, 58, 40, 30, 13, 13, 24], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ranked) + 1}"

    wb.save(path)
    return len(ranked), tier1


# --------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--window", default="24h", choices=["24h", "48h", "72h"])
    ap.add_argument("--out", default=None)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--config", default=None, help="path to a profile YAML file")
    args = ap.parse_args()

    apify_token = os.environ.get("APIFY_TOKEN")
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not apify_token:
        print("APIFY_TOKEN is not set", file=sys.stderr)
        return 1
    if not api_key and not args.dry_run:
        print("ANTHROPIC_API_KEY is not set", file=sys.stderr)
        return 1

    cfg = load_config(args.config)

    window_map = {"24h": ("r86400", 24.0), "48h": ("r172800", 48.0), "72h": ("r259200", 72.0)}
    apify_window, max_hours = window_map[args.window]

    jobs = collect(apify_token, apify_window, cfg["search"])
    if not jobs:
        print("No jobs collected. Both Apify passes failed.", file=sys.stderr)
        return 2
    print(f"[collect] {len(jobs)} unique postings")

    exclude_companies = {c.strip().lower() for c in cfg["search"]["exclude_companies"]}
    survivors, rejections = [], {}
    for j in jobs:
        reason = hard_filter(j, max_hours, exclude_companies)
        if reason:
            rejections[reason] = rejections.get(reason, 0) + 1
        else:
            survivors.append(j)

    print(f"[filter] {len(survivors)} survived hard filters")
    for reason, n in sorted(rejections.items(), key=lambda kv: -kv[1]):
        print(f"           {n:4d}  {reason}")

    if args.dry_run:
        for j in survivors[:40]:
            print(f"  {j.get('postedTime'):>16}  {j.get('companyName'):<28} {j.get('jobTitle')}")
        return 0

    system_prompt = build_scoring_system(cfg["profile"], cfg["rubric"])
    scored = score_all(api_key, system_prompt, survivors)

    tiers = cfg["tiers"]
    stamp = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d")
    out = args.out or f"job-matches-{stamp}.xlsx"
    rows, tier1 = write_xlsx(
        scored, out, cfg["referrals"],
        tiers["target_rows"], tiers["tier1_bar"], tiers["tier2_bar"],
    )

    print(f"\n[done] {out}")
    print(f"       {rows} rows written, {tier1} cleared the {tiers['tier1_bar']}% bar")
    if tier1 < tiers["target_rows"]:
        print(f"       Only {tier1} genuine matches today. The rest are labelled "
              f"Tier 2/3 rather than inflated. Run --window 48h to widen the pool.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
